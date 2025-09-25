import os
import json
import math
from typing import List, Optional, Dict, Any, Union

# Ensure headless backend for matplotlib image saving
os.environ.setdefault("MPLBACKEND", "Agg")

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
import pandas as pd

# Reuse existing computational functions from oob_eng
try:
    from oob_eng import (
        resource_path,
        load_execution_time,
        load_chart_information,
        preprocess_data,
        find_matching_file,
        determine_data_type,
        process_single_chart,
        plot_spc_chart,
        plot_weekly_spc_chart,
        calculate_cpk,
        ooc_calculator,
        review_ooc_results,
        discrete_oob_calculator,
        record_high_low_calculator,
    )
except Exception as e:
    # If import fails (e.g., missing PyQt), raise a clear error at startup
    raise RuntimeError(
        "Failed to import required functions from oob_eng.py. "
        "Please ensure all dependencies are installed. Error: " + str(e)
    )

# Tool Matching imports
try:
    from tool_matching_widget_osat import (
        analyze_tool_matching_data,
        get_k_value_headless,
        calculate_mean_index_headless,
        perform_statistical_test
    )
except Exception as e:
    print(f"Warning: Could not import tool matching functions: {e}")
    # Define dummy functions to prevent import errors
    def analyze_tool_matching_data(*args, **kwargs):
        raise HTTPException(status_code=500, detail="Tool matching functions not available")
    def get_k_value_headless(*args, **kwargs):
        return "N/A"
    def calculate_mean_index_headless(*args, **kwargs):
        return 0.0
    def perform_statistical_test(*args, **kwargs):
        return {"error": "Statistical test not available"}

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.transforms as mtransforms
from scipy import stats
import tempfile
import base64
from io import BytesIO
from datetime import datetime, date
import xlsxwriter
from typing import Tuple




class ProcessRequest(BaseModel):
    filepath: Optional[str] = Field(
        default=None,
        description="Path to All_Chart_Information.xlsx. Defaults to input/All_Chart_Information.xlsx",
    )
    raw_data_directory: Optional[str] = Field(
        default=None, description="Directory containing raw chart CSV files"
    )
    save_excel: bool = Field(default=True, description="Save Excel report with images")
    scale_factor: float = Field(default=0.3, description="Image scale factor in Excel")
    limit_charts: Optional[int] = Field(default=None, description="Limit number of charts to process")


class ProcessSummary(BaseModel):
    total_charts: int
    processed_charts: int
    skipped_charts: int
    excel_output: Optional[str] = None


class ResultItem(BaseModel):
    data_cnt: Optional[int] = None
    ooc_cnt: Optional[int] = None
    WE_Rule: Optional[str] = None
    OOB_Rule: Optional[str] = None
    data_type: Optional[str] = None
    Material_no: Optional[str] = None
    group_name: Optional[str] = None
    chart_name: Optional[str] = None
    chart_ID: Optional[str] = None
    Characteristics: Optional[str] = None
    USL: Optional[float] = None
    LSL: Optional[float] = None
    UCL: Optional[float] = None
    LCL: Optional[float] = None
    Target: Optional[float] = None
    Cpk: Optional[float] = None
    Resolution: Optional[float] = None
    HL_record_high_low: Optional[str] = None
    record_high: Optional[bool] = None
    record_low: Optional[bool] = None
    chart_path: Optional[str] = None
    weekly_chart_path: Optional[str] = None


class ProcessResponse(BaseModel):
    summary: ProcessSummary
    results: List[ResultItem]


# Tool Matching Models
class ToolMatchingRequest(BaseModel):
    filepath: str = Field(description="Path to CSV file with tool matching data")
    mean_index_threshold: float = Field(default=1.0, description="Mean index threshold")
    sigma_index_threshold: float = Field(default=2.0, description="Sigma index threshold") 
    use_statistical_test: bool = Field(default=False, description="Use statistical test instead of index")
    statistical_method: str = Field(default="unpaired", description="Statistical test method: unpaired or paired")
    alpha_level: float = Field(default=0.05, description="Significance level for statistical test")
    fill_sample_size: int = Field(default=5, description="Minimum sample size for analysis")
    filter_mode: str = Field(default="all_data", description="Data filter mode: all_data, specified_date, latest_data")
    base_date: Optional[str] = Field(default=None, description="Base date for filtered analysis (YYYY-MM-DD)")


class ToolMatchingSummary(BaseModel):
    total_groups: int
    abnormal_groups: int


class ToolMatchingResultItem(BaseModel):
    gname: str
    cname: str
    group: str
    group_all: str
    mean_index: Union[str, float]  # Could be numeric or text like "Significant" or "Insufficient Data"
    sigma_index: Union[str, float]  # Could be numeric or "Insufficient Data"
    k_value: Union[str, float]  # Could be numeric or "No Compare"
    mean: float
    std: float
    mean_median: Union[str, float]  # Could be numeric or "-"
    sigma_median: Union[str, float]  # Could be numeric or "-"
    n: int
    characteristic: str


class ToolMatchingResponse(BaseModel):
    summary: ToolMatchingSummary
    results: List[ToolMatchingResultItem]
    excel_output: Optional[str] = None


# SPC CPK Dashboard related models
class SPCCpkRequest(BaseModel):
    chart_excel_path: Optional[str] = Field(
        default=None,
        description="Path to All_Chart_Information.xlsx. Defaults to input/All_Chart_Information.xlsx"
    )
    raw_data_directory: Optional[str] = Field(
        default=None,
        description="Directory containing raw chart CSV files. Defaults to input/raw_charts"
    )
    start_date: Optional[date] = Field(
        default=None,
        description="Start date for analysis. If not provided, will use 3 months ago from end_date"
    )
    end_date: Optional[date] = Field(
        default=None,
        description="End date for analysis. If not provided, will use current date"
    )
    custom_mode: bool = Field(
        default=False,
        description="Whether to use custom time range mode"
    )
    selected_chart: Optional[str] = Field(
        default=None,
        description="Specific chart to analyze in format 'GroupName - ChartName'"
    )


class SPCCpkMetrics(BaseModel):
    cpk: Optional[float] = None
    cpk_l1: Optional[float] = None
    cpk_l2: Optional[float] = None
    custom_cpk: Optional[float] = None
    r1: Optional[float] = None
    r2: Optional[float] = None
    k_value: Optional[float] = None


class SPCChartInfo(BaseModel):
    group_name: str
    chart_name: str
    characteristics: str
    usl: Optional[float] = None
    lsl: Optional[float] = None
    target: Optional[float] = None
    metrics: SPCCpkMetrics
    chart_image: Optional[str] = None  # base64 encoded image
    mean_current: Optional[float] = None
    sigma_current: Optional[float] = None
    mean_last_month: Optional[float] = None
    sigma_last_month: Optional[float] = None
    mean_last2_month: Optional[float] = None
    sigma_last2_month: Optional[float] = None
    mean_all: Optional[float] = None
    sigma_all: Optional[float] = None


class SPCCpkResponse(BaseModel):
    charts: List[SPCChartInfo]
    summary: Dict[str, Any]
    excel_path: Optional[str] = None


class SplitRequest(BaseModel):
    mode: str = Field(
        description="Split mode: 'Type3_Horizontal' or 'Type2_Vertical'",
        pattern="^(Type3_Horizontal|Type2_Vertical)$",
    )
    input_files: List[str] = Field(description="List of CSV file paths to split")
    output_folder: Optional[str] = Field(
        default=None, 
        description="Base output folder; 'raw_charts' will be created inside. Defaults to 'input' if not specified"
    )


app = FastAPI(title="OOB/SPC FastAPI", version="1.0.0")

# 全域變數：記住最後一次分割的輸出資料夾路徑
_last_split_output_folder: Optional[str] = None


def _default_paths() -> Dict[str, str]:
    """Resolve default file paths similar to the original UI app."""
    filepath = resource_path("input/All_Chart_Information.xlsx")
    
    # 如果有最後分割的資料夾，優先使用；否則使用預設路徑
    if _last_split_output_folder and os.path.exists(_last_split_output_folder):
        raw_dir = _last_split_output_folder
    else:
        raw_dir = resource_path("input/raw_charts/")
    
    return {
        "filepath": filepath, 
        "raw_dir": raw_dir,
        # 新增 SPC CPK 所需的鍵名
        "chart_excel_path": filepath,
        "raw_data_directory": raw_dir
    }


def _read_csv_cached(cache: Dict[str, pd.DataFrame], filepath: str) -> Optional[pd.DataFrame]:
    try:
        if filepath not in cache:
            cache[filepath] = pd.read_csv(filepath)
        return cache[filepath].copy()
    except Exception as e:
        print(f"[Error] Failed to read CSV {filepath}: {e}")
        return None


def _preprocess_chart_types(all_charts_info: pd.DataFrame, raw_data_directory: str) -> Dict[str, str]:
    print("Preprocessing chart types...")
    chart_types: Dict[str, str] = {}
    processed_files = set()
    for _, info in all_charts_info.iterrows():
        group_name = str(info.get("GroupName", "Unknown"))
        chart_name = str(info.get("ChartName", "Unknown"))
        chart_key = f"{group_name}_{chart_name}"
        csv_path = find_matching_file(raw_data_directory, group_name, chart_name)
        if csv_path and os.path.exists(csv_path) and csv_path not in processed_files:
            try:
                head = pd.read_csv(csv_path, nrows=1000)
                if "point_val" in head.columns:
                    data_type = determine_data_type(head["point_val"].dropna())
                else:
                    data_type = "continuous"
                chart_types[chart_key] = data_type
                processed_files.add(csv_path)
            except Exception as e:
                print(f"  Type detect error for {chart_key}: {e}")
                chart_types[chart_key] = "continuous"
        else:
            chart_types[chart_key] = "continuous"
    print(f"Chart type preprocessing done: {len(chart_types)} entries")
    return chart_types


def _build_result_api(result: Dict[str, Any], violated_rules: Dict[str, bool], image_path: str, weekly_image_path: str) -> Dict[str, Any]:
    we_true_keys = [k for k, v in (violated_rules or {}).items() if v]
    result["WE_Rule"] = ", ".join(we_true_keys) if we_true_keys else "N/A"
    result["HL_WE"] = "HIGHLIGHT" if we_true_keys else "NO_HIGHLIGHT"

    OOB_KEYS = [
        "HL_P95_shift",
        "HL_P50_shift",
        "HL_P05_shift",
        "HL_sticking_shift",
        "HL_trending",
        "HL_high_OOC",
        "HL_record_high_low",
        "HL_category_LT_shift",
    ]
    oob_true_keys = [k for k in OOB_KEYS if result.get(k) == "HIGHLIGHT"]
    result["OOB_Rule"] = ", ".join(oob_true_keys) if oob_true_keys else "N/A"

    for k in OOB_KEYS:
        result.pop(k, None)
    result.pop("violated_rules", None)

    result["chart_path"] = image_path
    result["weekly_chart_path"] = weekly_image_path

    if not result.get("group_name"):
        result["group_name"] = str(result.get("GroupName", "N/A"))
    else:
        result["group_name"] = str(result["group_name"])
    if not result.get("chart_name"):
        result["chart_name"] = str(result.get("ChartName", "N/A"))
    else:
        result["chart_name"] = str(result["chart_name"])
    if "Cpk" not in result:
        result["Cpk"] = np.nan

    return result


def _process_discrete_chart_api(
    raw_df: pd.DataFrame,
    chart_info: Dict[str, Any],
    weekly_start_date: pd.Timestamp,
    weekly_end_date: pd.Timestamp,
    initial_baseline_start_date: pd.Timestamp,
    baseline_end_date: pd.Timestamp,
) -> Optional[Dict[str, Any]]:
    # Follow the logic of _process_discrete_chart in oob_eng, but without UI
    try:
        baseline_one_year = raw_df[(raw_df["point_time"] >= initial_baseline_start_date) & (raw_df["point_time"] <= baseline_end_date)].copy()
        baseline_count_one_year = len(baseline_one_year)
        baseline_insufficient = False

        if baseline_count_one_year < 10:
            actual_baseline_start = baseline_end_date - pd.Timedelta(days=365 * 2)
            baseline_two_year = raw_df[(raw_df["point_time"] >= actual_baseline_start) & (raw_df["point_time"] <= baseline_end_date)].copy()
            if len(baseline_two_year) < 10:
                baseline_insufficient = True
        else:
            actual_baseline_start = initial_baseline_start_date

        baseline_data = raw_df[(raw_df["point_time"] >= actual_baseline_start) & (raw_df["point_time"] <= baseline_end_date)].copy()
        weekly_data = raw_df[(raw_df["point_time"] >= weekly_start_date) & (raw_df["point_time"] <= weekly_end_date)].copy()

        baseline_empty = baseline_data.empty
        if weekly_data.empty:
            return None

        def calc_stats(df: pd.DataFrame) -> Dict[str, Any]:
            if df.shape[0] <= 1:
                sigma = 0.0
            else:
                sigma = df["point_val"].std()
            if np.isnan(sigma):
                sigma = 0.0
            return {
                "values": df["point_val"].values,
                "cnt": df.shape[0],
                "mean": df["point_val"].mean(),
                "sigma": sigma,
            }

        base_data_dict = calc_stats(baseline_data) if not baseline_empty else None
        weekly_data_dict = calc_stats(weekly_data)

        result: Dict[str, Any] = {
            "data_cnt": weekly_data_dict["cnt"],
            "ooc_cnt": 0,
            "WE_Rule": "",
            "OOB_Rule": "",
            "Material_no": chart_info.get("material_no", "N/A"),
            "group_name": str(chart_info.get("group_name", "N/A")),
            "chart_name": str(chart_info.get("chart_name", "N/A")),
            "chart_ID": chart_info.get("ChartID", "N/A"),
            "Characteristics": chart_info.get("Characteristics", "N/A"),
            "USL": chart_info.get("UCL", np.nan),
            "LSL": chart_info.get("LCL", np.nan),
            "UCL": chart_info.get("UCL", np.nan),
            "LCL": chart_info.get("LCL", np.nan),
            "Target": chart_info.get("Target", np.nan),
            "Resolution": chart_info.get("Resolution", np.nan),
            "baseline_insufficient": baseline_insufficient,
            "baseline_empty": baseline_empty,
            "data_type": "discrete",
        }

        if not baseline_insufficient and not baseline_empty:
            weekly_df = pd.DataFrame({"point_val": weekly_data["point_val"]})
            ooc_results = ooc_calculator(weekly_df, chart_info.get("UCL"), chart_info.get("LCL"))
            ooc_highlight = review_ooc_results(ooc_results[1], ooc_results[2])
            result["ooc_cnt"] = ooc_results[1]

            discrete_oob_result = discrete_oob_calculator(
                base_data_dict,
                weekly_data_dict,
                chart_info,
                raw_df,
                weekly_start_date,
                weekly_end_date,
                actual_baseline_start,
                baseline_end_date,
            )

            record_results = record_high_low_calculator(
                weekly_data["point_val"].values, baseline_data["point_val"].values
            )

            result.update(
                {
                    "HL_P95_shift": discrete_oob_result.get("HL_P95_shift", "NO_HIGHLIGHT"),
                    "HL_P50_shift": discrete_oob_result.get("HL_P50_shift", "NO_HIGHLIGHT"),
                    "HL_P05_shift": discrete_oob_result.get("HL_P05_shift", "NO_HIGHLIGHT"),
                    "HL_sticking_shift": discrete_oob_result.get("HL_sticking_shift", "NO_HIGHLIGHT"),
                    "HL_trending": discrete_oob_result.get("HL_trending", "NO_HIGHLIGHT"),
                    "HL_high_OOC": ooc_highlight,
                    "HL_category_LT_shift": discrete_oob_result.get("HL_category_LT_shift", "NO_HIGHLIGHT"),
                    "HL_record_high_low": record_results.get("highlight_status", "NO_HIGHLIGHT"),
                    "record_high": record_results.get("record_high", False),
                    "record_low": record_results.get("record_low", False),
                }
            )
        else:
            result.update(
                {
                    "HL_P95_shift": "NO_HIGHLIGHT",
                    "HL_P50_shift": "NO_HIGHLIGHT",
                    "HL_P05_shift": "NO_HIGHLIGHT",
                    "HL_sticking_shift": "NO_HIGHLIGHT",
                    "HL_trending": "NO_HIGHLIGHT",
                    "HL_high_OOC": "NO_HIGHLIGHT",
                    "HL_category_LT_shift": "NO_HIGHLIGHT",
                    "HL_record_high_low": "NO_HIGHLIGHT",
                    "record_high": False,
                    "record_low": False,
                }
            )

        return result
    except Exception as e:
        print(f"[Error] Discrete process failed: {e}")
        return None


def _analyze_chart_api(
    execution_time: Optional[pd.Timestamp], raw_df: pd.DataFrame, chart_info: Dict[str, Any]
) -> Optional[Dict[str, Any]]:
    # Determine WE rule list
    if "rule_list" not in chart_info or not chart_info.get("rule_list"):
        rule_list = []
        for rule in ["WE1", "WE2", "WE3", "WE4", "WE5", "WE6", "WE7", "WE8", "WE9", "WE10"]:
            if chart_info.get(rule, "N") == "Y":
                rule_list.append(rule)
        chart_info["rule_list"] = rule_list

    group_name = str(chart_info.get("group_name", chart_info.get("GroupName", "Unknown")))
    chart_name = str(chart_info.get("chart_name", chart_info.get("ChartName", "Unknown")))

    if "point_time" not in raw_df.columns or not pd.api.types.is_datetime64_any_dtype(raw_df["point_time"]):
        return None

    latest_raw_time = raw_df["point_time"].max()
    weekly_end_date = latest_raw_time if execution_time is None or pd.isna(execution_time) else execution_time
    if pd.isna(weekly_end_date):
        return None
    weekly_start_date = weekly_end_date - pd.Timedelta(days=6)
    baseline_end_date = weekly_start_date - pd.Timedelta(seconds=1)
    initial_baseline_start_date = baseline_end_date - pd.Timedelta(days=365)

    # Determine data type
    if raw_df is None or raw_df.empty or "point_val" not in raw_df.columns:
        data_type = "continuous"
    else:
        data_type = determine_data_type(raw_df["point_val"].dropna())
    chart_info["data_type"] = data_type

    if data_type == "discrete":
        result = _process_discrete_chart_api(
            raw_df,
            chart_info,
            weekly_start_date,
            weekly_end_date,
            initial_baseline_start_date,
            baseline_end_date,
        )
    else:
        result = process_single_chart(
            chart_info.copy(),
            raw_df,
            initial_baseline_start_date,
            baseline_end_date,
            weekly_start_date,
            weekly_end_date,
        )
        if result:
            result["data_type"] = "continuous"

    if result is None:
        return None

    # Generate images
    image_path, violated_rules = plot_spc_chart(raw_df, chart_info, weekly_start_date, weekly_end_date)
    weekly_image_path = plot_weekly_spc_chart(raw_df, chart_info, weekly_start_date, weekly_end_date)

    # Cpk on weekly data only
    weekly_data = raw_df[(raw_df["point_time"] >= weekly_start_date) & (raw_df["point_time"] <= weekly_end_date)].copy()
    cpk = calculate_cpk(weekly_data, chart_info)
    result["Cpk"] = cpk.get("Cpk", np.nan) if cpk else np.nan

    # Final normalize
    result = _build_result_api(result, violated_rules, image_path, weekly_image_path)
    return result


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/split-status")
def get_split_status() -> Dict[str, Any]:
    """獲取最後一次分割的狀態資訊"""
    global _last_split_output_folder
    
    if _last_split_output_folder and os.path.exists(_last_split_output_folder):
        # 計算資料夾中的 CSV 檔案數量
        try:
            csv_files = [f for f in os.listdir(_last_split_output_folder) if f.endswith('.csv')]
            file_count = len(csv_files)
        except Exception:
            file_count = 0
        
        return {
            "has_split_data": True,
            "split_folder": _last_split_output_folder,
            "csv_file_count": file_count,
            "folder_exists": True
        }
    else:
        return {
            "has_split_data": False,
            "split_folder": None,
            "csv_file_count": 0,
            "folder_exists": False
        }


@app.post("/process", response_model=ProcessResponse)
def process_charts_api(req: ProcessRequest) -> ProcessResponse:
    # Resolve paths
    defaults = _default_paths()
    filepath = req.filepath or defaults["filepath"]
    raw_dir = req.raw_data_directory or defaults["raw_dir"]

    if not os.path.isfile(filepath):
        raise HTTPException(status_code=400, detail=f"Excel file not found: {filepath}")
    if not os.path.isdir(raw_dir):
        os.makedirs(raw_dir, exist_ok=True)

    # Load chart info and execution time
    try:
        all_charts_info = load_chart_information(filepath)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read chart information: {e}")

    total_charts = len(all_charts_info)
    if req.limit_charts:
        all_charts_info = all_charts_info.head(req.limit_charts)
    exec_time = load_execution_time(filepath)

    # Preprocess types and CSV cache
    chart_types_cache = _preprocess_chart_types(all_charts_info, raw_dir)
    csv_cache: Dict[str, pd.DataFrame] = {}

    results: List[Dict[str, Any]] = []
    skipped = 0

    for _, chart_info_row in all_charts_info.iterrows():
        group_name = str(chart_info_row.get("GroupName"))
        chart_name = str(chart_info_row.get("ChartName"))
        chart_key = f"{group_name}_{chart_name}"
        csv_path = find_matching_file(raw_dir, group_name, chart_name)
        if not csv_path or not os.path.exists(csv_path):
            skipped += 1
            continue

        raw_df = _read_csv_cached(csv_cache, csv_path)
        if raw_df is None:
            skipped += 1
            continue

        # Prepare chart_info dict similar to preprocess_data usage
        chart_info = chart_info_row.copy()
        # Convert to dict first, then rename keys manually
        chart_info_dict = chart_info.to_dict()
        
        # Rename keys in dictionary
        if "Material_no" in chart_info_dict:
            chart_info_dict["material_no"] = chart_info_dict.pop("Material_no")
        if "GroupName" in chart_info_dict:
            chart_info_dict["group_name"] = chart_info_dict.pop("GroupName")  
        if "ChartName" in chart_info_dict:
            chart_info_dict["chart_name"] = chart_info_dict.pop("ChartName")
            
        chart_info = chart_info_dict

        # Ensure point_time is datetime
        if "point_time" in raw_df.columns:
            raw_df["point_time"] = pd.to_datetime(raw_df["point_time"], errors="coerce")
            raw_df.dropna(subset=["point_time"], inplace=True)

        # Preprocess
        is_ok, processed_df, updated_chart_info = preprocess_data(chart_info_row, raw_df)
        if not is_ok or processed_df is None or processed_df.empty:
            skipped += 1
            continue

        # Merge updated_chart_info onto dict for analysis/plotting
        chart_info.update(updated_chart_info.to_dict() if hasattr(updated_chart_info, "to_dict") else dict(updated_chart_info))

        # Determine type cache usage
        data_type = chart_types_cache.get(chart_key, "continuous")
        chart_info["data_type"] = data_type

        # Analyze
        result = _analyze_chart_api(exec_time, processed_df, chart_info)
        if result:
            results.append(result)
        else:
            skipped += 1

    processed = len(results)

    # Optionally save Excel using the function in oob_eng via a simple routine here
    excel_output = None
    if req.save_excel and results:
        try:
            results_df = pd.DataFrame(results)
            # Ensure expected columns exist
            expected_cols = [
                "data_cnt",
                "ooc_cnt",
                "WE_Rule",
                "OOB_Rule",
                "data_type",
                "Material_no",
                "group_name",
                "chart_name",
                "chart_ID",
                "Characteristics",
                "USL",
                "LSL",
                "UCL",
                "LCL",
                "Target",
                "Cpk",
                "Resolution",
                "HL_record_high_low",
                "record_high",
                "record_low",
                "chart_path",
                "weekly_chart_path",
            ]
            for col in expected_cols:
                if col not in results_df.columns:
                    results_df[col] = np.nan
            # Reorder
            results_df = results_df[[c for c in expected_cols if c in results_df.columns]]
            results_df = results_df.replace([np.nan, np.inf, -np.inf], "N/A")

            # Save with xlsxwriter and images similarly to oob_eng.save_results_to_excel
            from oob_eng import save_results_to_excel

            save_results_to_excel(results_df, scale_factor=req.scale_factor)
            excel_output = os.path.abspath("result_with_images.xlsx")
        except Exception as e:
            # Non-fatal: Excel export failed
            print(f"[Warn] Failed to save Excel: {e}")
            excel_output = None

    summary = ProcessSummary(
        total_charts=total_charts if not req.limit_charts else min(total_charts, req.limit_charts),
        processed_charts=processed,
        skipped_charts=skipped,
        excel_output=excel_output,
    )

    # Convert results to ResultItem, ensuring string types
    result_items = []
    for r in results:
        # Ensure string fields are strings
        if 'group_name' in r:
            r['group_name'] = str(r['group_name'])
        if 'chart_name' in r:
            r['chart_name'] = str(r['chart_name'])
        if 'Material_no' in r:
            r['Material_no'] = str(r['Material_no']) if r['Material_no'] is not None else None
        if 'chart_ID' in r:
            r['chart_ID'] = str(r['chart_ID']) if r['chart_ID'] is not None else None
        if 'Characteristics' in r:
            r['Characteristics'] = str(r['Characteristics']) if r['Characteristics'] is not None else None
        
        result_items.append(ResultItem(**r))
    return ProcessResponse(summary=summary, results=result_items)


def _sanitize_filename(name: str) -> str:
    invalid_chars = '<>:"/\\|?*\''
    for ch in invalid_chars:
        name = name.replace(ch, "")
    return name.strip()


def _read_csv_with_encoding_fallback(filepath: str, header_val=None) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "big5", "cp950", "latin1", "cp1252"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            return pd.read_csv(filepath, header=header_val, encoding=enc)
        except Exception as e:
            last_err = e
            continue
    raise ValueError(f"Failed to read {os.path.basename(filepath)} with common encodings: {last_err}")


def _split_type3_horizontal(input_path: str, final_output_folder: str) -> bool:
    try:
        df = _read_csv_with_encoding_fallback(input_path, header_val=None)
        new_columns: List[str] = []
        for col1, col2 in zip(df.iloc[0], df.iloc[1]):
            if pd.isna(col2):
                new_columns.append(str(col1))
            elif pd.isna(col1):
                new_columns.append(str(col2))
            else:
                new_columns.append(f"{col1}_{col2}")

        df = df.iloc[2:].copy()
        df.columns = new_columns

        chartname_col_name = None
        for col in df.columns:
            if "GroupName" in col and "ChartName" in col:
                chartname_col_name = col
                break
        if chartname_col_name is None:
            raise ValueError("Cannot find combined 'GroupName' and 'ChartName' header column")

        chartname_idx = df.columns.get_loc(chartname_col_name)
        universal_info_columns = df.columns[: chartname_idx + 1].tolist()
        chart_columns = df.columns[(chartname_idx + 1) :]

        for chart_col in chart_columns:
            temp_df = df[universal_info_columns].copy()
            temp_df["point_val"] = df[chart_col]
            if "_" in chart_col:
                groupname, chartname = chart_col.split("_", 1)
            else:
                groupname = ""
                chartname = chart_col
            temp_df["GroupName"] = groupname
            temp_df["ChartName"] = chartname
            if "point_time" in temp_df.columns:
                try:
                    temp_df["point_time"] = pd.to_datetime(temp_df["point_time"], errors="coerce")
                    # Windows-compatible strftime
                    temp_df["point_time"] = temp_df["point_time"].dt.strftime("%Y/%m/%d %H:%M")
                except Exception:
                    pass

            final_columns_order = ["GroupName", "ChartName", "point_time", "point_val"]
            for col in universal_info_columns:
                if col not in final_columns_order and col != chartname_col_name:
                    final_columns_order.append(col)
            existing_cols = [c for c in final_columns_order if c in temp_df.columns]
            temp_df = temp_df[existing_cols]

            safe_groupname = _sanitize_filename(str(groupname))
            safe_chartname = _sanitize_filename(str(chartname))
            output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
            if not temp_df.empty:
                temp_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"[Error] Type3 split failed for {os.path.basename(input_path)}: {e}")
        return False


def _split_type2_vertical(input_path: str, final_output_folder: str) -> bool:
    try:
        df = _read_csv_with_encoding_fallback(input_path, header_val="infer")
        required_cols = ["GroupName", "ChartName", "point_time", "point_val"]
        if not all(col in df.columns for col in required_cols):
            missing = [c for c in required_cols if c not in df.columns]
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        if "point_time" in df.columns:
            try:
                df["point_time"] = pd.to_datetime(df["point_time"], errors="coerce")
                df["point_time"] = df["point_time"].dt.strftime("%Y/%m/%d %H:%M")
            except Exception:
                pass
        uniq = df[["GroupName", "ChartName"]].drop_duplicates()
        for _, row in uniq.iterrows():
            groupname = row["GroupName"]
            chartname = row["ChartName"]
            temp_df = df[(df["GroupName"] == groupname) & (df["ChartName"] == chartname)].copy()
            other_cols = [c for c in temp_df.columns if c not in ["GroupName", "ChartName", "point_time", "point_val"]]
            final_cols = ["GroupName", "ChartName", "point_time", "point_val"] + other_cols
            existing = [c for c in final_cols if c in temp_df.columns]
            temp_df = temp_df[existing]
            safe_groupname = _sanitize_filename(str(groupname))
            safe_chartname = _sanitize_filename(str(chartname))
            output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
            if not temp_df.empty:
                temp_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"[Error] Type2 split failed for {os.path.basename(input_path)}: {e}")
        return False


@app.post("/split")
def split_csvs(req: SplitRequest) -> Dict[str, Any]:
    global _last_split_output_folder
    
    # 使用預設輸出資料夾如果未指定
    base_output_folder = req.output_folder or "input"
    final_output_folder = os.path.join(base_output_folder, "raw_charts")
    
    # 確保輸出資料夾存在
    os.makedirs(final_output_folder, exist_ok=True)
    
    successes = 0
    failures: List[str] = []
    for path in req.input_files:
        ok = False
        if req.mode == "Type3_Horizontal":
            ok = _split_type3_horizontal(path, final_output_folder)
        elif req.mode == "Type2_Vertical":
            ok = _split_type2_vertical(path, final_output_folder)
        if ok:
            successes += 1
        else:
            failures.append(os.path.basename(path))

    # 如果處理成功，記住這個資料夾路徑
    if successes > 0:
        _last_split_output_folder = os.path.abspath(final_output_folder)
        print(f"[Info] Split completed. Remembered output folder: {_last_split_output_folder}")

    return {
        "mode": req.mode,
        "output_folder": os.path.abspath(final_output_folder),
        "processed": successes,
        "failed": failures,
        "remembered_for_oob": successes > 0,
    }


@app.post("/tool-matching", response_model=ToolMatchingResponse)
def analyze_tool_matching(request: ToolMatchingRequest) -> ToolMatchingResponse:
    """
    執行 Tool Matching 分析
    - 讀取 CSV 檔案
    - 計算 Mean/Sigma Matching Index
    - 可選擇使用統計檢定方法
    - 生成圖表並輸出 Excel 報告
    """
    # 檢查檔案存在
    if not os.path.isfile(request.filepath):
        raise HTTPException(status_code=400, detail=f"CSV file not found: {request.filepath}")
    
    try:
        # 讀取 CSV 檔案
        df = pd.read_csv(request.filepath)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read CSV file: {e}")
    
    # 準備分析配置
    config = {
        "mean_index_threshold": request.mean_index_threshold,
        "sigma_index_threshold": request.sigma_index_threshold,
        "use_statistical_test": request.use_statistical_test,
        "statistical_method": request.statistical_method,
        "alpha_level": request.alpha_level,
        "fill_sample_size": request.fill_sample_size,
        "filter_mode": request.filter_mode,
        "base_date": request.base_date
    }
    
    try:
        # 執行分析並生成圖表和 Excel 報告
        analysis_result = _analyze_tool_matching_with_charts_and_excel(df, config, request.filepath)
        
        # 轉換結果格式
        result_items = []
        for _, row in analysis_result["results"].iterrows():
            # PyQt format: [gname, cname, group, group_all, mean_index, sigma_index, 
            #               k_value, mean, std, mean_median, sigma_median, n, characteristic]
            try:
                # 安全轉換數值，處理 inf 和 NaN 值
                def safe_float(val):
                    if pd.isna(val):
                        return 0.0
                    if isinstance(val, (int, float)):
                        if val == float('inf') or val == float('-inf'):
                            return 999999.0  # 用大數值代替 inf
                        if pd.isna(val):  # 檢查 NaN
                            return 0.0
                        return float(val)
                    return 0.0
                
                def safe_value(val, default='N/A'):
                    if pd.isna(val):
                        return default
                    if isinstance(val, float) and (val == float('inf') or val == float('-inf') or pd.isna(val)):
                        return 'Infinite' if val == float('inf') else 'Negative Infinite' if val == float('-inf') else default
                    return val
                
                result_items.append(ToolMatchingResultItem(
                    gname=str(row['gname']),
                    cname=str(row['cname']),
                    group=str(row['group']),
                    group_all=str(row['group_all']),
                    mean_index=safe_value(row['mean_index']),
                    sigma_index=safe_value(row['sigma_index']),
                    k_value=safe_value(row['k_value']),
                    mean=safe_float(row['mean']),
                    std=safe_float(row['std']),
                    mean_median=safe_value(row['mean_median']),
                    sigma_median=safe_value(row['sigma_median']),
                    n=int(row['n']) if pd.notna(row['n']) and isinstance(row['n'], (int, float)) else 0,
                    characteristic=str(row['characteristic'])
                ))
            except (KeyError, ValueError, TypeError) as e:
                print(f"Warning: Failed to parse result row: {row.to_dict()}, error: {e}")
                continue
        
        # 計算摘要統計
        total_groups = len(result_items)
        
        # 計算異常項目數（與 Streamlit 前端邏輯一致）
        abnormal_groups = 0
        for item in result_items:
            mean_index = item.mean_index
            sigma_index = item.sigma_index
            k_value = item.k_value
            
            # 檢查是否為資料不足
            is_data_insufficient = (
                mean_index == 'Insufficient Data' or 
                sigma_index == 'Insufficient Data' or 
                k_value == 'No Compare'
            )
            
            if not is_data_insufficient:
                is_abnormal = False
                
                # 檢查統計檢定顯著性
                if isinstance(mean_index, str) and ("Significant" in str(mean_index) or "ANOVA" in str(mean_index)):
                    if "No Significant" not in str(mean_index):
                        is_abnormal = True
                
                # 檢查指標門檻（使用用戶設定的參數）
                if not is_abnormal:
                    try:
                        mean_threshold = request.mean_index_threshold  # 使用用戶設定的門檻
                        sigma_threshold = request.sigma_index_threshold  # 使用用戶設定的 Sigma 門檻
                        
                        # 如果 sigma_threshold 是預設值 2.0，則使用 K 值作為門檻（與 PyQt 版本一致）
                        if sigma_threshold == 2.0 and k_value not in [None, '', 'No Compare']:
                            try:
                                sigma_threshold = float(k_value)
                            except (ValueError, TypeError):
                                pass
                        
                        # 檢查 Mean Index 異常（包括 Infinite）
                        mean_abn = False
                        if str(mean_index).lower() in ['infinite', 'inf', '-inf']:
                            mean_abn = True
                        elif isinstance(mean_index, (int, float)) and not (isinstance(mean_index, float) and (mean_index != mean_index)):
                            mean_abn = float(mean_index) >= mean_threshold
                        
                        # 檢查 Sigma Index 異常
                        sigma_abn = False
                        if str(sigma_index).lower() in ['infinite', 'inf', '-inf']:
                            sigma_abn = True
                        elif isinstance(sigma_index, (int, float)) and not (isinstance(sigma_index, float) and (sigma_index != sigma_index)):
                            sigma_abn = float(sigma_index) >= sigma_threshold
                        
                        if mean_abn or sigma_abn:
                            is_abnormal = True
                    except (ValueError, TypeError):
                        pass
                
                if is_abnormal:
                    abnormal_groups += 1
        
        summary = ToolMatchingSummary(
            total_groups=total_groups,
            abnormal_groups=abnormal_groups
        )
        
        return ToolMatchingResponse(
            summary=summary,
            results=result_items,
            excel_output=analysis_result.get("excel_output")
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")


def _analyze_tool_matching_with_charts_and_excel(df: pd.DataFrame, config: dict, source_path: str) -> dict:
    """
    完整的 Tool Matching 分析，包含圖表生成和 Excel 輸出
    """
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # 使用 headless backend
    import numpy as np
    import tempfile
    import os
    
    # 首先執行基本分析
    basic_result = analyze_tool_matching_data(df, config)
    
    # 準備圖表生成
    chart_figures = {}
    
    try:
        # 按 GroupName + ChartName 分組來生成圖表
        grouped_data = df.groupby(["GroupName", "ChartName"])
        
        for (group_name, chart_name), group_df in grouped_data:
            chart_key = (group_name, chart_name)
            
            try:
                # 生成 SPC 散點圖
                scatter_fig = _create_spc_chart(group_df, group_name, chart_name)
                
                # 生成盒鬚圖
                box_fig = _create_boxplot_chart(group_df, group_name, chart_name)
                
                chart_figures[chart_key] = {
                    'scatter': scatter_fig,
                    'box': box_fig
                }
                
            except Exception as e:
                print(f"Warning: Failed to generate charts for {group_name}/{chart_name}: {e}")
                continue
    
    except Exception as e:
        print(f"Warning: Chart generation failed: {e}")
    
    # 生成 Excel 報告
    excel_output = None
    try:
        excel_output = _export_tool_matching_to_excel(basic_result["results"], chart_figures, source_path)
    except Exception as e:
        print(f"Warning: Excel export failed: {e}")
    
    # 清理圖表物件以釋放記憶體
    for figures in chart_figures.values():
        try:
            if 'scatter' in figures:
                plt.close(figures['scatter'])
            if 'box' in figures:
                plt.close(figures['box'])
        except:
            pass
    
    return {
        "summary": basic_result["summary"],
        "results": basic_result["results"],
        "excel_output": excel_output
    }


def _create_spc_chart(group_df: pd.DataFrame, group_name: str, chart_name: str):
    """生成 SPC 散點圖（與 PyQt 版本一致）"""
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    import numpy as np
    from matplotlib import cm
    
    fig, ax = plt.subplots(figsize=(7, 4.5))  # 調整尺寸為較小的長方形
    
    # 依 matching_group 字母順序排序
    unique_groups = sorted(group_df["matching_group"].unique(), key=lambda x: str(x))
    labels = [str(mg) for mg in unique_groups]
    
    # 檢查是否有數據可供繪圖
    if group_df.empty or not any(len(grp["point_val"]) > 0 for _, grp in group_df.groupby("matching_group")):
        print(f"[WARNING] Skipping chart creation for {group_name} - {chart_name} due to empty data.")
        return None
    
    # 為不同的組設置顏色
    colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))
    
    # 確保 point_time 是 datetime 格式
    if 'point_time' in group_df.columns:
        group_df = group_df.copy()
        group_df['point_time'] = pd.to_datetime(group_df['point_time'], errors='coerce')
    
    # 為每個群組繪製數據點，按時間順序連線
    x_position = 0
    for i, mg in enumerate(unique_groups):
        group_data = group_df[group_df["matching_group"] == mg]
        if 'point_time' in group_df.columns:
            group_data = group_data.sort_values("point_time")
        
        if not group_data.empty:
            # 為每個群組創建連續的x位置
            x_vals = np.arange(x_position, x_position + len(group_data))
            y_vals = group_data["point_val"].values
            
            # 繪製數據點
            ax.scatter(x_vals, y_vals, color=colors[i], alpha=0.8, s=40, label=f'{mg}', zorder=3)
            
            # 連接同組內的點
            ax.plot(x_vals, y_vals, color=colors[i], alpha=0.5, linewidth=1, zorder=2)
            
            # 在群組間添加分隔線
            if i < len(unique_groups) - 1:  # 不在最後一組後面加線
                separator_x = x_position + len(group_data) - 0.5
                ax.axvline(x=separator_x, color='gray', linestyle='-', alpha=0.3, zorder=1)
            
            x_position += len(group_data)
    
    # 設置圖表樣式
    ax.set_title(f"SPC Chart: {group_name} - {chart_name}", fontsize=10)
    ax.set_xlabel("Sample Sequence (Grouped by Matching Group)")
    ax.set_ylabel("Point Value")
    ax.grid(True, linestyle='--', alpha=0.3, zorder=0)
    
    # 添加群組標籤在x軸上
    if unique_groups:
        group_positions = []
        x_pos = 0
        for mg in unique_groups:
            group_size = len(group_df[group_df["matching_group"] == mg])
            group_positions.append(x_pos + group_size/2 - 0.5)
            x_pos += group_size
        
        # 設置x軸刻度和標籤
        ax.set_xticks(group_positions)
        ax.set_xticklabels(labels, rotation=0, ha='center')
        
        # 添加次要刻度顯示樣本序號
        ax.tick_params(axis='x', which='minor', bottom=True, top=False)
    
    # 調整圖例位置
    ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
    plt.tight_layout()
    
    return fig


def _create_boxplot_chart(group_df: pd.DataFrame, group_name: str, chart_name: str):
    """生成盒鬚圖（與 PyQt 版本一致）"""
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib import cm
    
    fig, ax = plt.subplots(figsize=(7, 4.5))  # 調整尺寸為較小的長方形
    
    # 依 matching_group 字母順序排序
    unique_groups = sorted(group_df["matching_group"].unique(), key=lambda x: str(x))
    labels = [str(mg) for mg in unique_groups]
    
    # 依排序後 unique_groups 組裝 box_data，確保顏色/label/資料一致
    box_data = [group_df[group_df["matching_group"] == mg]["point_val"].values for mg in unique_groups]
    group_stats = group_df.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count'])
    
    # 為不同的組設置顏色
    colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))
    
    if box_data:
        bp = ax.boxplot(box_data, labels=labels, patch_artist=True, widths=0.6)
        for patch, color in zip(bp['boxes'], colors):
            patch.set_facecolor(color)
        
        # legend 也照 unique_groups 順序
        legend_labels = [
            f"{label}: μ={group_stats.loc[mg, 'mean']:.2f}, σ={group_stats.loc[mg, 'std']:.2f}, n={int(group_stats.loc[mg, 'count'])}"
            for label, mg in zip(labels, unique_groups)
        ]
        ax.legend([bp["boxes"][i] for i in range(len(labels))], legend_labels, loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
    
    ax.set_title(f"Boxplot: {group_name} - {chart_name}", fontsize=10)
    ax.set_xlabel("Matching Group")
    ax.set_ylabel("Point Value")
    ax.grid(True, linestyle='--', alpha=0.6)
    fig.subplots_adjust(right=0.7)
    plt.tight_layout()
    
    return fig


def _export_tool_matching_to_excel(results_df: pd.DataFrame, chart_figures: dict, source_path: str) -> str:
    """匯出 Tool Matching 結果到 Excel，返回暫存檔案路徑（與 PyQt 版本格式一致）"""
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        import tempfile
        
        # 轉換結果格式以符合 PyQt 版本
        # PyQt 格式: [gname, cname, group, group_all, mean_index, sigma_index, k_value, mean, std, mean_median, sigma_median, n, characteristic]
        # Excel 格式: SPC_Chart, BoxPlot, Need_matching, AbnormalType, GroupName, ChartName, matching_group, mean_matching_index, sigma_matching_index, K, mean, sigma, mean_median, sigma_median, samplesize, characteristic
        
        excel_data = []
        for _, row in results_df.iterrows():
            # 判斷是否異常（簡化邏輯，可依需求調整）
            mean_index = row['mean_index']
            sigma_index = row['sigma_index']
            k_value = row['k_value']
            
            need_matching = False
            abnormal_type = ""
            
            # 判斷異常條件
            if isinstance(mean_index, (int, float)) and isinstance(sigma_index, (int, float)) and isinstance(k_value, (int, float)):
                if abs(mean_index) > 2.0:  # Mean 異常
                    need_matching = True
                    abnormal_type = "Mean"
                elif sigma_index > k_value:  # Sigma 異常
                    need_matching = True
                    abnormal_type = "Sigma"
            elif "Significant" in str(mean_index) and "No Significant" not in str(mean_index):
                need_matching = True
                abnormal_type = "Mean"
            
            # 處理特殊值
            def format_value(val):
                if pd.isna(val):
                    return ""
                if isinstance(val, float) and (val == float('inf') or val == float('-inf')):
                    return "inf" if val == float('inf') else "-inf"
                return val
            
            excel_data.append({
                'SPC_Chart': "",  # 圖片將在後續插入
                'BoxPlot': "",    # 圖片將在後續插入
                'Need_matching': need_matching,
                'AbnormalType': abnormal_type,
                'GroupName': str(row['gname']),
                'ChartName': str(row['cname']),
                'matching_group': str(row['group']),
                'mean_matching_index': format_value(row['mean_index']),
                'sigma_matching_index': format_value(row['sigma_index']),
                'K': format_value(row['k_value']),
                'mean': format_value(row['mean']),
                'sigma': format_value(row['std']),  # std -> sigma
                'mean_median': format_value(row['mean_median']),
                'sigma_median': format_value(row['sigma_median']),
                'samplesize': int(row['n']) if pd.notna(row['n']) else 0,
                'characteristic': str(row['characteristic'])
            })
        
        # 建立 DataFrame
        df = pd.DataFrame(excel_data)
        
        # 生成暫存檔案路徑（不儲存到原始目錄）
        temp_dir = "temp_uploads"
        os.makedirs(temp_dir, exist_ok=True)
        file_name = os.path.splitext(os.path.basename(source_path))[0]
        output_path = os.path.join(temp_dir, f"{file_name}_matching_results.xlsx")
        
        # 創建圖片暫存目錄
        img_temp_dir = tempfile.mkdtemp()
        
        try:
            # 寫入 Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Tool Matching Results', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Tool Matching Results']
                
                # 設定標題格式
                header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                header_fill = openpyxl.styles.PatternFill(start_color="344CB7", end_color="344CB7", fill_type="solid")
                header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 設定欄寬
                worksheet.column_dimensions['A'].width = 70  # SPC Chart
                worksheet.column_dimensions['B'].width = 70  # Box Plot
                
                # 異常行格式
                abnormal_fill = openpyxl.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                
                # 嵌入圖表
                img_display_width, img_display_height = 450, 250
                
                for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
                    is_abnormal = row_data["Need_matching"]
                    
                    # 異常行標記
                    if is_abnormal:
                        for cell in worksheet[row_idx]:
                            cell.fill = abnormal_fill
                    
                    # 嵌入圖表
                    group_name = str(row_data["GroupName"])
                    chart_name = str(row_data["ChartName"])
                    chart_key = (group_name, chart_name)
                    
                    if chart_key in chart_figures:
                        try:
                            # SPC 圖
                            scatter_fig = chart_figures[chart_key]['scatter']
                            temp_scatter_path = os.path.join(img_temp_dir, f"spc_{group_name}_{chart_name}_{row_idx}.png")
                            scatter_fig.savefig(temp_scatter_path, format='png', bbox_inches='tight', dpi=100)
                            
                            scatter_img = XLImage(temp_scatter_path)
                            scatter_img.width = img_display_width
                            scatter_img.height = img_display_height
                            worksheet.add_image(scatter_img, f"A{row_idx}")
                            
                            # 盒鬚圖
                            box_fig = chart_figures[chart_key]['box']
                            temp_box_path = os.path.join(img_temp_dir, f"box_{group_name}_{chart_name}_{row_idx}.png")
                            box_fig.savefig(temp_box_path, format='png', bbox_inches='tight', dpi=100)
                            
                            box_img = XLImage(temp_box_path)
                            box_img.width = img_display_width
                            box_img.height = img_display_height
                            worksheet.add_image(box_img, f"B{row_idx}")
                            
                        except Exception as e:
                            print(f"Warning: Failed to embed charts for row {row_idx}: {e}")
                            worksheet.cell(row=row_idx, column=1).value = "Chart failed to load"
                            worksheet.cell(row=row_idx, column=2).value = "Chart failed to load"
                
                # 設定行高以容納圖片
                for row in range(2, len(df) + 2):
                    worksheet.row_dimensions[row].height = img_display_height * 0.75
        
        finally:
            # 清理圖片暫存檔案
            try:
                import shutil
                shutil.rmtree(img_temp_dir)
            except:
                pass
        
        print(f"Excel report saved to temp location: {output_path}")
        return output_path
        
    except Exception as e:
        print(f"Excel export failed: {e}")
        return None

# Optional root endpoint info
@app.get("/")
def root() -> Dict[str, Any]:
    defaults = _default_paths()
    return {
        "message": "OOB/SPC FastAPI is running",
        "defaults": defaults,
        "usage": {
            "process": "POST /process with {filepath, raw_data_directory, save_excel}",
            "split": "POST /split with {mode: Type3_Horizontal|Type2_Vertical, input_files, output_folder}",
            "tool-matching": "POST /tool-matching with {filepath, mean_index_threshold, sigma_index_threshold, ...}",
            "spc-cpk": "POST /spc-cpk with {chart_excel_path, raw_data_directory, start_date, end_date, custom_mode, selected_chart}",
            "health": "GET /health",
        },
    }

def calculate_cpk_dashboard(raw_df: pd.DataFrame, chart_info: dict) -> dict:
    """計算CPK指標，支援三種特性類型"""
    if raw_df.empty:
        return {'Cpk': None}
    
    mean = raw_df['point_val'].mean()
    std = raw_df['point_val'].std()
    characteristic = chart_info.get('Characteristics', '')
    usl = chart_info.get('USL', None)
    lsl = chart_info.get('LSL', None)
    
    cpk = None
    if std > 0:
        if characteristic == 'Nominal':
            if usl is not None and lsl is not None:
                cpu = (usl - mean) / (3 * std)
                cpl = (mean - lsl) / (3 * std)
                cpk = min(cpu, cpl)
        elif characteristic in ['Smaller', 'Sigma']:
            if usl is not None:
                cpk = (usl - mean) / (3 * std)
        elif characteristic == 'Bigger':
            if lsl is not None:
                cpk = (mean - lsl) / (3 * std)
    
    if cpk is not None:
        cpk = round(cpk, 3)
    return {'Cpk': cpk}


def compute_cpk_windows(raw_df: pd.DataFrame, chart_info: dict, end_time: pd.Timestamp) -> dict:
    """計算多時間窗口的CPK值，同時計算各窗口 mean / sigma 以便直接輸出到 Excel。

    回傳欄位：
      - Cpk, Cpk_last_month, Cpk_last2_month
      - mean_current, sigma_current
      - mean_last_month, sigma_last_month
      - mean_last2_month, sigma_last2_month
    """
    result = {
        'Cpk': None,
        'Cpk_last_month': None,
        'Cpk_last2_month': None,
        'mean_current': None,
        'sigma_current': None,
        'mean_last_month': None,
        'sigma_last_month': None,
        'mean_last2_month': None,
        'sigma_last2_month': None,
    }

    if raw_df is None or raw_df.empty:
        return result

    if 'point_time' not in raw_df.columns:
        # 沒有時間欄位：只能算全期間 (視為 current)
        sub_cpk = calculate_cpk_dashboard(raw_df, chart_info)['Cpk']
        result['Cpk'] = sub_cpk
        result['mean_current'] = raw_df['point_val'].mean()
        result['sigma_current'] = raw_df['point_val'].std()
        return result

    df = raw_df.copy()
    try:
        df['point_time'] = pd.to_datetime(df['point_time'])
    except Exception:
        # 轉換失敗直接返回全期間統計
        result['Cpk'] = calculate_cpk_dashboard(df, chart_info)['Cpk']
        result['mean_current'] = df['point_val'].mean()
        result['sigma_current'] = df['point_val'].std()
        return result

    # 限制截止時間
    df = df[df['point_time'] <= end_time]
    if df.empty:
        return result

    start1 = end_time - pd.DateOffset(months=1)
    start2 = end_time - pd.DateOffset(months=2)
    start3 = end_time - pd.DateOffset(months=3)

    # 原本條件為 (startX, end] 左開右閉，這裡保留；若需要含 start 請再調整 >=。
    mask1 = (df['point_time'] > start1) & (df['point_time'] <= end_time)
    mask2 = (df['point_time'] > start2) & (df['point_time'] <= start1)
    mask3 = (df['point_time'] > start3) & (df['point_time'] <= start2)

    # Current (L0)
    if mask1.any():
        seg = df[mask1]
        result['Cpk'] = calculate_cpk_dashboard(seg, chart_info)['Cpk']
        result['mean_current'] = seg['point_val'].mean()
        result['sigma_current'] = seg['point_val'].std()
    # Last month (L1)
    if mask2.any():
        seg = df[mask2]
        result['Cpk_last_month'] = calculate_cpk_dashboard(seg, chart_info)['Cpk']
        result['mean_last_month'] = seg['point_val'].mean()
        result['sigma_last_month'] = seg['point_val'].std()
    # Last 2 month (L2)
    if mask3.any():
        seg = df[mask3]
        result['Cpk_last2_month'] = calculate_cpk_dashboard(seg, chart_info)['Cpk']
        result['mean_last2_month'] = seg['point_val'].mean()
        result['sigma_last2_month'] = seg['point_val'].std()

    # 數值清理：避免 Inf / NaN 造成 JSON 序列化錯誤
    for k, v in list(result.items()):
        if isinstance(v, float):
            if math.isnan(v) or math.isinf(v):
                result[k] = None
    return result


def generate_spc_chart_base64(raw_df: pd.DataFrame, chart_info: dict, 
                             start_date: Optional[date] = None, 
                             end_date: Optional[date] = None, 
                             custom_mode: bool = False) -> str:
    """生成SPC圖表並返回base64編碼"""
    fig = plt.figure(figsize=(12, 6))
    
    # 創建GridSpec布局
    import matplotlib.gridspec as gridspec
    gs = gridspec.GridSpec(2, 2, width_ratios=[3, 1], height_ratios=[1, 1], 
                          hspace=0.3, wspace=0.25)
    
    # 主SPC圖
    ax_main = fig.add_subplot(gs[:, 0])
    ax_box = fig.add_subplot(gs[0, 1])  # Box Plot
    ax_qq = fig.add_subplot(gs[1, 1])   # QQ Plot
    
    if raw_df is None or raw_df.empty:
        ax_main.text(0.5, 0.5, "No Data", ha='center', va='center', transform=ax_main.transAxes)
        ax_box.text(0.5, 0.5, "No Data", ha='center', va='center', transform=ax_box.transAxes)
        ax_qq.text(0.5, 0.5, "No Data", ha='center', va='center', transform=ax_qq.transAxes)
    else:
        plot_df = raw_df.copy()
        
        # 日期過濾
        if start_date and end_date and 'point_time' in plot_df.columns:
            try:
                plot_df['point_time'] = pd.to_datetime(plot_df['point_time'])
                start_ts = pd.to_datetime(start_date)
                end_ts = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(milliseconds=1)
                filtered = plot_df[(plot_df['point_time'] >= start_ts) & (plot_df['point_time'] <= end_ts)]
                if not filtered.empty:
                    plot_df = filtered
            except Exception:
                pass
        
        if not plot_df.empty:
            # 繪製主SPC圖
            _draw_main_spc_chart_api(ax_main, plot_df, chart_info, start_date, end_date, custom_mode)
            # 繪製Box Plot
            _draw_box_plot_api(ax_box, plot_df, chart_info)
            # 繪製QQ Plot
            _draw_qq_plot_api(ax_qq, plot_df, chart_info)
    
    # 設置標題
    group_name = chart_info.get('GroupName', '')
    chart_name = chart_info.get('ChartName', '')
    characteristics = chart_info.get('Characteristics', '')
    ax_main.set_title(f"{group_name}@{chart_name}@{characteristics}", pad=18, fontsize=12)
    
    # Box Plot標題
    if raw_df is not None and 'EQP_id' in raw_df.columns and not raw_df['EQP_id'].isna().all():
        ax_box.set_title("Box Plot (by EQP_id)", fontsize=10)
    else:
        ax_box.set_title("Box Plot", fontsize=10)
    
    ax_qq.set_title("Q-Q Plot", fontsize=10)
    
    fig.tight_layout()
    
    # 轉換為base64
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close(fig)
    
    return image_base64


def _draw_main_spc_chart_api(ax, plot_df, chart_info, start_date, end_date, custom_mode):
    """繪製主要的SPC控制圖"""
    y = plot_df['point_val'].values
    x = range(1, len(y) + 1)
    
    # 等距模式處理時間排序
    if 'point_time' in plot_df.columns:
        try:
            plot_df = plot_df.sort_values('point_time').reset_index(drop=True)
            y = plot_df['point_val'].values
        except Exception:
            pass
    
    # 標示時間區間
    if 'point_time' in plot_df.columns and not plot_df.empty:
        try:
            times = pd.to_datetime(plot_df['point_time']).to_numpy()
            tmin, tmax = times.min(), times.max()
            
            if custom_mode and start_date and end_date:
                # 自訂模式：顯示自訂範圍
                start_time = pd.to_datetime(start_date)
                end_time = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(milliseconds=1)
                windows = [(start_time, end_time, 'Custom', '#dbeafe')]
            else:
                # 原本邏輯：三個月窗口
                end_sel = pd.to_datetime(end_date) if end_date else pd.Timestamp(tmax)
                if end_sel > pd.Timestamp(tmax):
                    end_sel = pd.Timestamp(tmax)
                start1 = end_sel - pd.DateOffset(months=1)
                start2 = end_sel - pd.DateOffset(months=2)
                start3 = end_sel - pd.DateOffset(months=3)
                windows = [
                    (start1, end_sel, 'L0', '#dbeafe'),
                    (start2, start1, 'L1', '#fef9c3'),
                    (start3, start2, 'L2', '#ede9fe'),
                ]
            
            text_trans = mtransforms.blended_transform_factory(ax.transData, ax.transAxes)
            n = len(times)
            
            def t2ix_left(t):
                return float(np.searchsorted(times, np.datetime64(t), side='left')) + 0.5
            def t2ix_right(t):
                return float(np.searchsorted(times, np.datetime64(t), side='right')) + 0.5
            
            x_min, x_max = 0.5, n + 0.5
            for s, e, lab, col in windows:
                s_clip = max(pd.Timestamp(s), pd.Timestamp(tmin))
                e_clip = min(pd.Timestamp(e), pd.Timestamp(tmax))
                if e_clip <= s_clip:
                    continue
                xl = max(x_min, t2ix_left(s_clip))
                xr = min(x_max, t2ix_right(e_clip))
                if xr <= xl:
                    continue
                ax.axvspan(xl, xr, color=col, alpha=0.25, zorder=0)
                x_center = (xl + xr) / 2.0
                ax.text(x_center, 1.04, lab, transform=text_trans, ha='center', va='top', 
                       fontsize=8, color='#374151', alpha=0.9)
        except Exception:
            pass
    
    # 繪製資料點和線
    ax.plot(x, y, linestyle='-', marker='o', color='#2563eb', markersize=4, linewidth=1.0)
    
    # 取得控制限
    usl = chart_info.get('USL', None)
    lsl = chart_info.get('LSL', None)
    target = None
    for key_t in ['Target', 'TARGET', 'TargetValue', '中心線', 'Center']:
        if key_t in chart_info and pd.notna(chart_info.get(key_t)):
            target = chart_info[key_t]
            break
    
    mean_val = float(np.mean(y)) if len(y) else None
    
    # 標示超規點
    if usl is not None:
        ax.scatter([xi for xi, yi in zip(x, y) if yi > usl], 
                  [yi for yi in y if yi > usl], 
                  color='#dc2626', s=25, zorder=5)
    if lsl is not None:
        ax.scatter([xi for xi, yi in zip(x, y) if yi < lsl], 
                  [yi for yi in y if yi < lsl], 
                  color='#dc2626', marker='s', s=25, zorder=5)
    
    # 計算Y軸範圍
    extra_vals = [v for v in [usl, lsl, target, mean_val] 
                  if v is not None and not (isinstance(v, float) and np.isnan(v))]
    if len(y) > 0:
        ymin_sel = float(np.min(y))
        ymax_sel = float(np.max(y))
    else:
        ymin_sel, ymax_sel = (0.0, 1.0)
    if extra_vals:
        ymin_sel = min(ymin_sel, min(extra_vals))
        ymax_sel = max(ymax_sel, max(extra_vals))
    rng = ymax_sel - ymin_sel
    margin = 0.05 * rng if rng > 0 else 1.0
    ax.set_ylim(ymin_sel - margin, ymax_sel + margin)
    
    # 繪製控制線
    trans = mtransforms.blended_transform_factory(ax.transAxes, ax.transData)
    
    def segment_with_label(val, name, color, va='center'):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return
        x0, x1 = 0.0, 0.96
        ax.plot([x0, x1], [val, val], transform=trans, color=color, linestyle='--', linewidth=1.0)
        ax.text(x1, val, name, transform=trans, color=color, va=va, ha='left', fontsize=8)
    
    segment_with_label(usl, 'USL', '#ef4444', va='center')
    segment_with_label(lsl, 'LSL', '#ef4444', va='center')
    segment_with_label(target, 'Target', '#f59e0b', va='center')
    segment_with_label(mean_val, 'Mean', '#16a34a', va='center')
    
    # X軸刻度
    if 'point_time' in plot_df.columns and not plot_df.empty:
        times = plot_df['point_time'].tolist()
        total = len(times)
        if total <= 8:
            tick_idx = list(range(1, total + 1))
        else:
            step = max(1, total // 6)
            tick_idx = list(range(1, total + 1, step))
            if tick_idx[-1] != total:
                tick_idx.append(total)
        labels = [times[i-1].strftime('%Y-%m-%d') for i in tick_idx]
        ax.set_xticks(tick_idx)
        ax.set_xticklabels(labels, rotation=30, ha='right', fontsize=8)
    
    ax.grid(True, linestyle=':', linewidth=0.6, alpha=0.5)


def _draw_box_plot_api(ax, plot_df, chart_info):
    """繪製箱型圖"""
    if plot_df.empty:
        ax.text(0.5, 0.5, "No Data", ha='center', va='center', transform=ax.transAxes)
        return
    
    # 檢查是否有EQP_id欄位
    if 'EQP_id' in plot_df.columns and not plot_df['EQP_id'].isna().all():
        # 按EQP_id分組
        grouped = plot_df.groupby('EQP_id')
        eqp_ids = list(grouped.groups.keys())
        
        box_data = []
        labels = []
        
        for eqp_id in sorted(eqp_ids):
            group_data = grouped.get_group(eqp_id)['point_val'].values
            if len(group_data) > 0:
                box_data.append(group_data)
                labels.append(str(eqp_id))
        
        if len(box_data) == 0:
            ax.text(0.5, 0.5, "No Valid Data", ha='center', va='center', transform=ax.transAxes)
            return
        
        box_plot = ax.boxplot(box_data, patch_artist=True, notch=False)
        
        # 設定顏色
        colors = ['#87CEEB', '#98FB98', '#FFB6C1', '#F0E68C', '#DDA0DD', '#F5DEB3', '#B0E0E6']
        for i, patch in enumerate(box_plot['boxes']):
            patch.set_facecolor(colors[i % len(colors)])
            patch.set_alpha(0.8)
        
        ax.set_xticklabels(labels, rotation=0, ha='center', fontsize=9)
        ax.set_xlabel('')
    else:
        # 單一box plot
        y = plot_df['point_val'].values
        if len(y) == 0:
            ax.text(0.5, 0.5, "No Data", ha='center', va='center', transform=ax.transAxes)
            return
        
        box_plot = ax.boxplot(y, patch_artist=True, notch=False)
        box_plot['boxes'][0].set_facecolor('#87CEEB')
        box_plot['boxes'][0].set_alpha(0.8)
        
        ax.set_xticks([])
        ax.set_xlabel('')


def _draw_qq_plot_api(ax, plot_df, chart_info):
    """繪製Q-Q圖"""
    y = plot_df['point_val'].values
    if len(y) == 0:
        ax.text(0.5, 0.5, "No Data", ha='center', va='center', transform=ax.transAxes)
        return
    
    try:
        # 計算Q-Q plot資料
        (osm, osr), (slope, intercept, r) = stats.probplot(y, dist="norm", plot=None)
        
        # 繪製資料點
        ax.scatter(osm, osr, alpha=0.7, color='blue', s=20)
        
        # 繪製理論線
        line_x = np.array([osm.min(), osm.max()])
        line_y = slope * line_x + intercept
        ax.plot(line_x, line_y, 'r-', linewidth=1.5, alpha=0.8, label=f'R²={r**2:.3f}')
        
        ax.set_xlabel('Theoretical Quantiles', fontsize=8)
        ax.set_ylabel('Sample Quantiles', fontsize=8)
        ax.grid(True, linestyle=':', linewidth=0.6, alpha=0.3)
        ax.tick_params(axis='both', which='major', labelsize=8)
        ax.legend(fontsize=7, loc='lower right')
        
    except Exception as e:
        ax.text(0.5, 0.5, f"Calculation Error:\n{str(e)}", ha='center', va='center',
               transform=ax.transAxes, fontsize=8)
@app.post("/spc-cpk", response_model=SPCCpkResponse)
def analyze_spc_cpk(request: SPCCpkRequest) -> SPCCpkResponse:
    """
    執行 SPC CPK Dashboard 分析
    - 讀取圖表資訊和原始資料
    - 計算 CPK 指標 (當月/L1/L2/全期間)
    - 生成 SPC 控制圖、Box Plot、Q-Q Plot
    - 可匯出 Excel 詳細報告
    """
    # 設定預設路徑 - 使用統一的 _default_paths 函數
    defaults = _default_paths()
    chart_excel_path = request.chart_excel_path or defaults["chart_excel_path"]
    raw_data_directory = request.raw_data_directory or defaults["raw_data_directory"]
    
    # 檢查檔案存在
    if not os.path.isfile(chart_excel_path):
        raise HTTPException(status_code=400, detail=f"Chart information file not found: {chart_excel_path}")
    
    if not os.path.isdir(raw_data_directory):
        raise HTTPException(status_code=400, detail=f"Raw data directory not found: {raw_data_directory}")
    
    try:
        # 載入圖表資訊
        all_charts_info = load_chart_information(chart_excel_path)
        if all_charts_info is None or all_charts_info.empty:
            raise HTTPException(status_code=400, detail="No chart information loaded")
        
        # 設定分析日期範圍
        end_date = request.end_date or date.today()
        if request.start_date:
            start_date = request.start_date
        else:
            # 預設為結束日期往前3個月
            start_date = end_date - pd.DateOffset(months=3)
            start_date = start_date.date()
        
        # 載入原始資料和計算 CPK
        raw_charts_dict = {}
        chart_results = []
        
        for _, chart_info in all_charts_info.iterrows():
            group_name = str(chart_info['GroupName'])
            chart_name = str(chart_info['ChartName'])
            characteristics = str(chart_info.get('Characteristics', ''))
            
            # 如果指定了特定圖表，只處理該圖表
            if request.selected_chart:
                chart_display_name = f"{group_name} - {chart_name}"
                if chart_display_name != request.selected_chart:
                    continue
            
            # 尋找對應的原始資料檔案
            raw_path = find_matching_file(raw_data_directory, group_name, chart_name)
            
            if raw_path and os.path.exists(raw_path):
                try:
                    raw_df = pd.read_csv(raw_path)
                    
                    # 資料預處理：移除超出控制限的資料
                    usl = chart_info.get('USL', None)
                    lsl = chart_info.get('LSL', None)
                    if usl is not None and lsl is not None:
                        raw_df = raw_df[(raw_df['point_val'] <= usl) & (raw_df['point_val'] >= lsl)]
                    elif usl is not None:
                        raw_df = raw_df[raw_df['point_val'] <= usl]
                    elif lsl is not None:
                        raw_df = raw_df[raw_df['point_val'] >= lsl]
                    
                    raw_charts_dict[(group_name, chart_name)] = raw_df
                    
                    # 計算 CPK 指標
                    if request.custom_mode and request.start_date and request.end_date:
                        # 自訂時間模式：只計算指定範圍的 CPK
                        cpk_res = _compute_cpk_custom_range(raw_df, chart_info.to_dict(), 
                                                          pd.to_datetime(start_date), 
                                                          pd.to_datetime(end_date))
                        all_data_cpk = calculate_cpk_dashboard(raw_df, chart_info.to_dict())['Cpk']
                        metrics = SPCCpkMetrics(
                            cpk=cpk_res.get('Cpk'),
                            cpk_l1=None,  # 自訂模式不顯示 L1/L2
                            cpk_l2=None,
                            custom_cpk=all_data_cpk,
                            r1=None,
                            r2=None,
                            k_value=_calculate_k_value(raw_df, chart_info.to_dict(), start_date, end_date, request.custom_mode)
                        )
                    else:
                        # 標準模式：計算三個月窗口
                        end_time = pd.to_datetime(end_date)
                        if 'point_time' in raw_df.columns:
                            raw_df_local = raw_df.copy()
                            raw_df_local['point_time'] = pd.to_datetime(raw_df_local['point_time'])
                            latest = raw_df_local['point_time'].max()
                            if end_time > latest:
                                end_time = latest
                        
                        cpk_res = compute_cpk_windows(raw_df, chart_info.to_dict(), end_time)
                        all_data_cpk = calculate_cpk_dashboard(raw_df, chart_info.to_dict())['Cpk']
                        
                        # 計算 R1, R2 衰退率
                        cpk = cpk_res.get('Cpk')
                        l1 = cpk_res.get('Cpk_last_month')
                        l2 = cpk_res.get('Cpk_last2_month')
                        
                        r1 = r2 = None
                        if cpk is not None and l1 is not None and l1 != 0 and cpk <= l1:
                            r1 = (1 - (cpk / l1)) * 100
                        if cpk is not None and l1 is not None and l2 is not None and l2 != 0 and cpk <= l1 <= l2:
                            r2 = (1 - (cpk / l2)) * 100
                        
                        metrics = SPCCpkMetrics(
                            cpk=cpk,
                            cpk_l1=l1,
                            cpk_l2=l2,
                            custom_cpk=all_data_cpk,
                            r1=r1,
                            r2=r2,
                            k_value=_calculate_k_value(raw_df, chart_info.to_dict(), start_date, end_date, request.custom_mode)
                        )

                    # 先取得原本統計（保留既有邏輯），再用 cpk_res 的 mean/sigma 覆蓋缺失值
                    mean_stats = _calculate_period_statistics(raw_df, end_date, request.custom_mode, start_date)
                    # 如果 compute_cpk_windows 有提供數值而 mean_stats 為 None，則填入
                    for k_src, k_dst in [
                        ('mean_current', 'mean_current'),
                        ('sigma_current', 'sigma_current'),
                        ('mean_last_month', 'mean_last_month'),
                        ('sigma_last_month', 'sigma_last_month'),
                        ('mean_last2_month', 'mean_last2_month'),
                        ('sigma_last2_month', 'sigma_last2_month'),
                    ]:
                        if mean_stats.get(k_dst) is None and cpk_res.get(k_src) is not None:
                            mean_stats[k_dst] = cpk_res[k_src]
                    
                    # 生成圖表
                    chart_image = generate_spc_chart_base64(raw_df, chart_info.to_dict(), 
                                                          start_date, end_date, request.custom_mode)
                    
                    chart_info_obj = SPCChartInfo(
                        group_name=group_name,
                        chart_name=chart_name,
                        characteristics=characteristics,
                        usl=chart_info.get('USL'),
                        lsl=chart_info.get('LSL'),
                        target=_get_target_value(chart_info),
                        metrics=metrics,
                        chart_image=chart_image,
                        **mean_stats
                    )
                    
                    chart_results.append(chart_info_obj)
                    
                except Exception as e:
                    print(f"Error processing chart {group_name}/{chart_name}: {e}")
                    continue
            else:
                print(f"Raw data file not found for {group_name}/{chart_name}")
                continue
        
        # 生成摘要統計（擴充）
        cpk_values = [c.metrics.cpk for c in chart_results if c.metrics.cpk is not None]
        all_cpk_values = [c.metrics.custom_cpk for c in chart_results if c.metrics.custom_cpk is not None]
        mean_curr_list = [c.mean_current for c in chart_results if c.mean_current is not None]
        total_points_current = 0
        total_points_all = 0
        # 嘗試估算每個 current 視窗點數（需要 point_time，因此粗略重算）
        # 為避免昂貴計算，僅在有 point_time 時做彙總
        # （簡化：不額外讀檔，這裡略過細緻點數統計，只提供已計算欄位的聚合）
        summary = {
            "total_charts": len(chart_results),
            "charts_with_cpk": len(cpk_values),
            "avg_cpk": round(float(np.mean(cpk_values)), 4) if cpk_values else None,
            "median_cpk": round(float(np.median(cpk_values)), 4) if cpk_values else None,
            "avg_all_cpk": round(float(np.mean(all_cpk_values)), 4) if all_cpk_values else None,
            "median_all_cpk": round(float(np.median(all_cpk_values)), 4) if all_cpk_values else None,
            "charts_with_mean_current": len(mean_curr_list),
            "custom_mode": request.custom_mode,
            "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # 可選：生成 Excel 報告
        excel_path = None
        if len(chart_results) > 0:
            try:
                excel_path = _export_spc_cpk_to_excel(chart_results, summary, start_date, end_date)
            except Exception as e:
                print(f"Warning: Excel export failed: {e}")
        
        # 最終輸出前再全域清理一次圖表中的 Inf/NaN，避免 JSON 失敗
        def _sanitize_number(x):
            if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
                return None
            return x
        for ch in chart_results:
            ch.mean_current = _sanitize_number(ch.mean_current)
            ch.sigma_current = _sanitize_number(ch.sigma_current)
            ch.mean_last_month = _sanitize_number(ch.mean_last_month)
            ch.sigma_last_month = _sanitize_number(ch.sigma_last_month)
            ch.mean_last2_month = _sanitize_number(ch.mean_last2_month)
            ch.sigma_last2_month = _sanitize_number(ch.sigma_last2_month)
            ch.mean_all = _sanitize_number(ch.mean_all)
            ch.sigma_all = _sanitize_number(ch.sigma_all)
            # metrics 中的值
            ch.metrics.cpk = _sanitize_number(ch.metrics.cpk)
            ch.metrics.cpk_l1 = _sanitize_number(ch.metrics.cpk_l1)
            ch.metrics.cpk_l2 = _sanitize_number(ch.metrics.cpk_l2)
            ch.metrics.custom_cpk = _sanitize_number(ch.metrics.custom_cpk)
            ch.metrics.r1 = _sanitize_number(ch.metrics.r1)
            ch.metrics.r2 = _sanitize_number(ch.metrics.r2)
            ch.metrics.k_value = _sanitize_number(ch.metrics.k_value)

        return SPCCpkResponse(
            charts=chart_results,
            summary=summary,
            excel_path=excel_path
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SPC CPK analysis failed: {str(e)}")


def _compute_cpk_custom_range(raw_df: pd.DataFrame, chart_info: dict, start_time: pd.Timestamp, end_time: pd.Timestamp) -> dict:
    """根據自訂的起始和結束時間計算 Cpk"""
    result = {'Cpk': None, 'Cpk_last_month': None, 'Cpk_last2_month': None}
    
    if raw_df is None or raw_df.empty:
        return result
    
    if 'point_time' not in raw_df.columns:
        result['Cpk'] = calculate_cpk_dashboard(raw_df, chart_info)['Cpk']
        return result
    
    df = raw_df.copy()
    df['point_time'] = pd.to_datetime(df['point_time'])
    
    # 篩選自訂範圍的資料
    filtered_df = df[(df['point_time'] >= start_time) & (df['point_time'] <= end_time)]
    
    if filtered_df.empty:
        return result
    
    # 只計算自訂範圍的 Cpk（當月）
    result['Cpk'] = calculate_cpk_dashboard(filtered_df, chart_info)['Cpk']
    
    return result


def _calculate_k_value(raw_df: pd.DataFrame, chart_info: dict, start_date: date, end_date: date, custom_mode: bool) -> Optional[float]:
    """計算 K 值（偏移度指標）"""
    try:
        usl = chart_info.get('USL')
        lsl = chart_info.get('LSL')
        target = _get_target_value(chart_info)
        
        if target is None or usl is None or lsl is None:
            return None
        
        # 計算均值（根據模式選擇範圍）
        if custom_mode and 'point_time' in raw_df.columns:
            # 自訂模式：使用指定範圍內的資料
            start_ts = pd.to_datetime(start_date)
            end_ts = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(milliseconds=1)
            filtered_df = raw_df[(pd.to_datetime(raw_df['point_time']) >= start_ts) & 
                               (pd.to_datetime(raw_df['point_time']) <= end_ts)]
            if not filtered_df.empty:
                mean_val = filtered_df['point_val'].mean()
            else:
                mean_val = raw_df['point_val'].mean()
        else:
            mean_val = raw_df['point_val'].mean()
        
        rng = (usl - lsl) / 2 if (usl - lsl) != 0 else None
        if mean_val is not None and rng:
            k_val = abs(mean_val - target) / rng
            return round(k_val, 3)
        
    except Exception:
        pass
    
    return None


def _get_target_value(chart_info) -> Optional[float]:
    """取得目標值（支援多種欄位名稱）"""
    target_keys = ['Target', 'TARGET', 'TargetValue', '中心線', 'Center']
    for key in target_keys:
        if key in chart_info and pd.notna(chart_info.get(key)):
            return chart_info[key]
    return None


def _calculate_period_statistics(raw_df: pd.DataFrame, end_date: date, custom_mode: bool, start_date: Optional[date] = None) -> dict:
    """計算各時段的統計值"""
    stats = {
        'mean_current': None,
        'sigma_current': None,
        'mean_last_month': None,
        'sigma_last_month': None,
        'mean_last2_month': None,
        'sigma_last2_month': None,
        'mean_all': None,
        'sigma_all': None
    }
    
    if raw_df is None or raw_df.empty:
        return stats
    
    # 全期間統計
    stats['mean_all'] = raw_df['point_val'].mean()
    stats['sigma_all'] = raw_df['point_val'].std()
    
    if 'point_time' not in raw_df.columns:
        return stats
    
    try:
        df = raw_df.copy()
        df['point_time'] = pd.to_datetime(df['point_time'])
        end_time = pd.to_datetime(end_date)
        
        if custom_mode and start_date:
            # 自訂模式：只計算當月統計
            start_time = pd.to_datetime(start_date)
            end_time = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(milliseconds=1)
            current_df = df[(df['point_time'] >= start_time) & (df['point_time'] <= end_time)]
            if not current_df.empty:
                stats['mean_current'] = current_df['point_val'].mean()
                stats['sigma_current'] = current_df['point_val'].std()
        else:
            # 標準模式：計算三個月窗口
            start1 = end_time - pd.DateOffset(months=1)
            start2 = end_time - pd.DateOffset(months=2)
            start3 = end_time - pd.DateOffset(months=3)
            
            # 當月
            current_df = df[(df['point_time'] > start1) & (df['point_time'] <= end_time)]
            if not current_df.empty:
                stats['mean_current'] = current_df['point_val'].mean()
                stats['sigma_current'] = current_df['point_val'].std()
            
            # 上月
            last_month_df = df[(df['point_time'] > start2) & (df['point_time'] <= start1)]
            if not last_month_df.empty:
                stats['mean_last_month'] = last_month_df['point_val'].mean()
                stats['sigma_last_month'] = last_month_df['point_val'].std()
            
            # 上上月
            last2_month_df = df[(df['point_time'] > start3) & (df['point_time'] <= start2)]
            if not last2_month_df.empty:
                stats['mean_last2_month'] = last2_month_df['point_val'].mean()
                stats['sigma_last2_month'] = last2_month_df['point_val'].std()
    
    except Exception as e:
        print(f"Warning: Period statistics calculation failed: {e}")
    
    return stats


def _export_spc_cpk_to_excel(chart_results: List[SPCChartInfo], summary: dict, start_date: date, end_date: date) -> str:
    """匯出 SPC CPK 結果到 Excel，包含圖片"""
    try:
        import xlsxwriter
        import tempfile
        
        # 生成暫存檔案路徑
        temp_dir = "temp_uploads"
        os.makedirs(temp_dir, exist_ok=True)
        output_path = os.path.join(temp_dir, f"spc_cpk_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        # 準備 Excel 資料
        excel_data = []
        chart_images = []
        
        for chart in chart_results:
            # 處理圖片
            if chart.chart_image:
                # 儲存 base64 圖片到暫存檔案
                temp_img_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                temp_img_file.close()
                
                try:
                    import base64
                    img_data = base64.b64decode(chart.chart_image)
                    with open(temp_img_file.name, 'wb') as f:
                        f.write(img_data)
                    chart_images.append(temp_img_file.name)
                except Exception as e:
                    print(f"Warning: Failed to save chart image: {e}")
                    chart_images.append(None)
            else:
                chart_images.append(None)
            
            # 準備資料行
            excel_data.append({
                'ChartImage': '',  # 佔位，稍後插入圖片
                'ChartKey': f"{chart.group_name}@{chart.chart_name}@{chart.characteristics}",
                'GroupName': chart.group_name,
                'ChartName': chart.chart_name,
                'Characteristics': chart.characteristics,
                'USL': chart.usl,
                'LSL': chart.lsl,
                'Target': chart.target,
                'K': chart.metrics.k_value,
                'Cpk_Curr': chart.metrics.cpk,
                'Cpk_L1': chart.metrics.cpk_l1,
                'Cpk_L2': chart.metrics.cpk_l2,
                'Custom_Cpk': chart.metrics.custom_cpk,
                'R1(%)': chart.metrics.r1,
                'R2(%)': chart.metrics.r2,
                'Mean_Curr': chart.mean_current,
                'Sigma_CurrentMonth': chart.sigma_current,
                'Mean_LastMonth': chart.mean_last_month,
                'Sigma_LastMonth': chart.sigma_last_month,
                'Mean_Last2Month': chart.mean_last2_month,
                'Sigma_Last2Month': chart.sigma_last2_month,
                'Mean_All': chart.mean_all,
                'Sigma_All': chart.sigma_all
            })
        
        # 建立 DataFrame
        df = pd.DataFrame(excel_data)
        
        # 欄位順序
        columns = ['ChartImage'] + [c for c in df.columns if c != 'ChartImage']
        
        # 建立 Excel 檔案
        workbook = xlsxwriter.Workbook(output_path)
        worksheet = workbook.add_worksheet()
        
        # 設定欄寬
        worksheet.set_column(0, 0, 100)  # 圖片欄位寬度調整為 120
        for i in range(1, len(columns)):
            worksheet.set_column(i, i, 15)  # 其他欄位寬度調整為 15
        
        # 標題格式
        bold = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter'})
        cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
        
        # 寫入標題
        for col_idx, col_name in enumerate(columns):
            worksheet.write(0, col_idx, col_name, bold)

        
        for row_idx, (row_data, img_path) in enumerate(zip(df.to_dict('records'), chart_images), 1):
            # 插入圖片
            if img_path and os.path.exists(img_path):
                worksheet.set_row(row_idx, 200)  # 行高調整為 200
                worksheet.insert_image(row_idx, 0, img_path, {
                    'x_scale': 0.6,  # 水平保持原始大小
                    'y_scale': 0.4,  # 垂直縮放為 60%，讓圖片變扁
                    'object_position': 1,
                    'y_offset': 10
                })
            
            # 寫入其他欄位
            for col_idx, col_name in enumerate(columns[1:], 1):
                val = row_data.get(col_name, '')
                # 處理 NaN/Inf 問題
                if val is None:
                    val = ''
                elif isinstance(val, float):
                    if math.isnan(val) or math.isinf(val):
                        val = 'N/A'
                worksheet.write(row_idx, col_idx, val, cell_format)
        
        workbook.close()
        
        # 清理暫存圖片檔案
        for img_path in chart_images:
            if img_path and os.path.exists(img_path):
                try:
                    os.unlink(img_path)
                except:
                    pass
        
        print(f"SPC CPK Excel report saved to: {output_path}")
        return output_path
        
    except Exception as e:
        print(f"Excel export failed: {e}")
        return None


# Optional root endpoint info
@app.get("/")
def root() -> Dict[str, Any]:
    defaults = _default_paths()
    return {
        "message": "OOB/SPC FastAPI is running",
        "defaults": defaults,
        "usage": {
            "process": "POST /process with {filepath, raw_data_directory, save_excel}",
            "split": "POST /split with {mode: Type3_Horizontal|Type2_Vertical, input_files, output_folder}",
            "tool-matching": "POST /tool-matching with {filepath, mean_index_threshold, sigma_index_threshold, ...}",
            "spc-cpk": "POST /spc-cpk with {chart_excel_path, raw_data_directory, start_date, end_date, custom_mode, selected_chart}",
            "health": "GET /health",
        },
    }
